import cv2
import numpy as np
from scipy.spatial import distance as dist
import time
import matplotlib.pyplot as plt
import csv
from scipy.spatial import distance as dist
import datetime
import pandas as pd
from keras.preprocessing.image import load_img
from keras.preprocessing.image import img_to_array
from mrcnn.config import Config
from mrcnn.model import MaskRCNN
from matplotlib import pyplot
from matplotlib.patches import Rectangle
from PyQt5 import QtCore, QtGui, QtWidgets
from gui import Ui_MainWindow
import sys
import win32api,win32process,win32con
import imutils
import threading
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter




class TestConfig(Config):
    #NAME = "test"
    #NAME ="kangaroo_cfg"
    NAME ="man_mov"
    GPU_COUNT = 1
    IMAGES_PER_GPU = 1
    #NUM_CLASSES = 1 + 11 ###
    NUM_CLASSES = 1 + 12

global rcnn
global class_id
global lookForOcr
global timeOver
global valCsv
global rcnn
global ui
global MainWindow

#pid = win32api.GetCurrentProcessId()
#print(pid,"<----Process ID--------------")
#handle = win32api.OpenProcess(win32con.PROCESS_ALL_ACCESS, True, pid)
#win32process.SetPriorityClass(handle, win32process.HIGH_PRIORITY_CLASS)

lookForOcr = False
timeOver = 0

# define the model
rcnn = MaskRCNN(mode='inference', model_dir='./', config=TestConfig())
# load coco model weights
rcnn.load_weights('model_newData/ocr_0036_1_june_2020.h5', by_name=True)
# load class
#class_id = ["BG","0","1","2","3","4","5","6","7","8","A","C"] ###
class_id = ["BG","0","1","2","3","4","5","6","7","8","9","A","C"]


rcnn.keras_model._make_predict_function()

def createNewFileImage(xlxsName):
    wb = openpyxl.Workbook()
    wb.save(xlxsName)
    
def writeCSVImage(inputImage,data1):
    filename="trailSaveImg_3.xlsx"
    width = 100
    height = 70
     
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        print("-----no file foud-----")
        createNewFileImage(filename)
        wb = load_workbook(filename)
    ws = wb.worksheets[0]

    latestCol = ws.max_column + 1
    colmnLetter = get_column_letter(latestCol)
    print(latestCol ,colmnLetter, "<-- col anchors.............")

    ws.row_dimensions[2].height = height
    ws.column_dimensions[colmnLetter].width = width
    for row, entry in enumerate(data1,start=3):
        ws.cell(row=row, column=latestCol, value=entry)

    img = Image(inputImage)
    img.width = 650 
    img.height = 110
    print(img.width)
    print(img.height)
    #img.resize((width,height),Image.NEAREST)
    
    ws.add_image(img,anchor=colmnLetter+str(1))
    wb.save(filename)

def drawBBox(boxes_list, score,idx):
    boxWithCharector = []
    for i,box in enumerate(boxes_list):
        if score[i] > .8:
            
            boxList = list(box)
            boxList.append(class_id[idx[i]])
            boxList.append(score[i])
            
            boxWithCharector.append(tuple(boxList))
            
            #y1, x1, y2, x2 = box
            #print(box)            
            
    return boxWithCharector
def takeSecond(i):
    return i[1]

def predOCR(img):
    global rcnn
    global lookForOcr
    global timeOver
    global valCsv
    global ui
    global imageName
    
    img = cv2.cvtColor(img,cv2.COLOR_BGR2RGB)
    img_a = img_to_array(img)
    # make prediction
    print("----detect--------")
    #rcnn._  #model._make_predict_function()
    #with graph.as_default():
    stt = time.time()
    results = rcnn.detect([img_a], verbose=0)
    # visualize the results
    print("**************************************************>", time.time()-stt,"sec")
    print("----look for results--------")
    cv2.imwrite("images_for_trail/"+time.strftime("%Y-%m-%d(%H_%M_%S)")+".jpg",img)
    #if len(results[0]['rois'] > 5):
    boxWithCharector = drawBBox(results[0]['rois'],results[0]["scores"],results[0]["class_ids"])
    if len(boxWithCharector) > 5:
        print("-----char available-------")
        detectedOCR = ''
        detectedOCROnscreen = ''
        boxWithCharector.sort(key=takeSecond)
        for box in boxWithCharector:
            y1, x1, y2, x2,className, classScore = box
            img = cv2.rectangle(img,(x1, y1), (x2, y2), (0,255,0),3)
            img = cv2.putText(img,"{}".format(className),(x1,y1), cv2.FONT_HERSHEY_SIMPLEX, .75 , (0,0,255), 2 , cv2.LINE_AA)
            #print(class_id[idx[i]])
            detectedOCR += className
            detectedOCROnscreen += className+"->"+str(classScore)+str(" ")
        
        '''if len(valCsv)>30:
                valCsv[1] = detectedOCR
                
                writeCSV(valCsv)
                valCsv = []'''
        
        imageName = time.strftime("%Y-%m-%d(%H_%M_%S)")        
        valCsv[1] = detectedOCR
        valCsv[2] = detectedOCROnscreen
        writeCSV(valCsv) 
        
        
        #lookForOcr = False
        timeOver = 0
        ui.updateOCR(detectedOCR)
        # save image with ocr printed
        
        cv2.imwrite("images_ocr/"+str(imageName)+".jpg",img)
        print("++++++++OCR and write CSV=>",detectedOCROnscreen)
        print("----------------------OCR=>",detectedOCR)
        print(lookForOcr,'<---- look for new slab')
        print(timeOver,'<----- Time',imageName)
        writeCSVImage("images_ocr/"+str(imageName)+".jpg", valCsv)#imageName
        valCsv = []
        t4 = threading.Thread(target=waitForSlabToPass)
        t4.start()
        t4.join()

        #time.sleep(20)
        #lookForOcr = False
        #print('------------')
        return img
    return img

def waitForSlabToPass():
    global lookForOcr
    time.sleep(20)
    lookForOcr = False
    return None

def ocr(ocrCameraLink):
    #c = np.load("ocrPoints.npy")
    global lookForOcr
    global timeOver
    global valCsv
    global ui
    #global MainWindow
    #global Ui_MainWindow
    
    lookForOcr = False
    print(lookForOcr)


    #blk = cv2.imread("blank.jpg",0)
    #ocr = cv2.VideoCapture("SLAB ID OCR & SMS Slab/SLAB ID OCR_2020-04-07_192853_195300.mov") #  1
    #ocr = cv2.VideoCapture('stable_camera/set_1/0 - 2020-05-19 12-40-52-663.mov') #'http://192.168.0.3:8080'
    #ocr = cv2.VideoCapture("C:/Users/CCTV/Desktop/Videos/1015.mov")
    ocr = cv2.VideoCapture(ocrCameraLink)
    ocr.set(cv2.CAP_PROP_BUFFERSIZE, 0)
    cnt = 0
    #ocr.set(cv2.CAP_PROP_POS_FRAMES,1500+1500+750)
    while True:
        #start = time.time()
        _, ocr_image = ocr.read()
        
        if _ == False:
            print("OCR camera Video could not be fetched --- Please close browser and try again thread-2")
            ocr = cv2.VideoCapture(ocrCameraLink)
            ocr.set(cv2.CAP_PROP_BUFFERSIZE, 0)
            continue
        img = ocr_image[680:900,550:1450]
        img = imutils.rotate_bound(img,-.81)
        img_dim = cv2.resize(ocr_image ,(int(1200),int(700)))
        #cv2.imshow("crop", img)
        #cv2.imshow("OcrFull",img_dim1)
        img_dim = cv2.cvtColor(img_dim, cv2.COLOR_BGR2RGB)
        img_dim = np.require(img_dim,np.uint8,'C')
        ui.showImg2(img_dim)

        if lookForOcr == True and (time.time()-timeOver) < 150.0:
                       
            predOCR(img)
            ocr = cv2.VideoCapture(ocrCameraLink)
            ocr.set(cv2.CAP_PROP_BUFFERSIZE, 0)
            
        elif ((time.time()-timeOver) >= 160.0 and (time.time()-timeOver) < 380.) :
            #timeOver = 0
            ocr = cv2.VideoCapture(ocrCameraLink)
            ocr.set(cv2.CAP_PROP_BUFFERSIZE, 0)
            #None
            print("inside else time for OCR is over")          
            valCsv[1] = "NA----"
                
            writeCSV(valCsv)
            writeCSVImage(str("blankSlab")+".jpg", valCsv)#imageName
            valCsv = []
            lookForOcr = False
            timeOver = 0
            ui.updateOCR("No Marking..")
            print(lookForOcr,'<---- Lock')
            print(timeOver,'<----- Time')
            print("Write CSV -------with no OCR or NO Marking")
            print("<----look for new slab")

            #cv2.imshow("h",h)
            #cv2.imshow('thresh',thresh)
            #cv2.imshow('H space', h)
        
        #print(time.time()-start,"ms--------")
        #cv2.imshow("crop", img)
        #cv2.imshow("OcrFull",img_dim1)

        #img_dim = cv2.cvtColor(img_dim, cv2.COLOR_BGR2RGB)
        #img_dim = np.require(img_dim,np.uint8,'C')
        #ui.showImg2(img_dim)
        continue
        '''key = cv2.waitKey(1) & 0xFF
        if key == ord('q'):
            break         
        
    ocr.release()
    cv2.destroyAllWindows()'''
def uiocrView(ocrCameraLink):
    global ui
    ocr = cv2.VideoCapture(ocrCameraLink)
    ocr.set(cv2.CAP_PROP_BUFFERSIZE, 0)
    cnt = 0
    #ocr.set(cv2.CAP_PROP_POS_FRAMES,1500+1500+750)
    while True:
        #start = time.time()
        _, ocr_image = ocr.read()
        
        if _ == False:
            print("OCR camera Video could not be fetched --- Please close browser and try again -- thread-3")
            ocr = cv2.VideoCapture(ocrCameraLink)
            ocr.set(cv2.CAP_PROP_BUFFERSIZE, 0)
            continue
        img_ocr = cv2.resize(ocr_image ,(int(1200),int(700)))
        img_ocr = cv2.cvtColor(img_ocr, cv2.COLOR_BGR2RGB)
        img_ocr = np.require(img_ocr,np.uint8,'C')
        ui.showImg2(img_ocr)
        continue

def mmPerPix(img_dim,thresh,alar,pixelsPerMetric):
   
    leftSide = np.nonzero(thresh[alar[1],])[0][0] 
    rightSide = np.nonzero(thresh[alar[3],])[0][-1]

    mmPpix = ((1/pixelsPerMetric)* (rightSide - leftSide)) -65 #* 0.9864

    img_dim  = cv2.line(img_dim ,(leftSide ,alar[1]),(rightSide ,alar[3]),(0,255,0),4)

    img_dim = cv2.circle(img_dim,(leftSide,alar[1]) , 1, 0, 5)
    img_dim = cv2.circle(img_dim,(rightSide,alar[3]) , 1, 0, 5)

    img_dim = cv2.putText(img_dim,"{} mm".format(mmPpix),(0,330), cv2.FONT_HERSHEY_SIMPLEX, 2 , (200,255,155), 2 , cv2.LINE_AA)

    return int(mmPpix),img_dim

def threshOP(img_dim):

    
    #b,g,r = cv2.split(img_dim.copy())
    b = img_dim[:,:,0]
    yuv = cv2.cvtColor(img_dim,cv2.COLOR_BGR2YUV)
    #y,u,vb = cv2.split(yuv)
    y = yuv[:,:,0]
    #randv = cv2.addWeighted(y, .5, b, 2. ,0.0)
    randv = cv2.addWeighted(y, .5, b, 2. ,0.0)
    dst = cv2.addWeighted(y, 0.005, randv, 1.4, 0.0)

    blur = cv2.GaussianBlur(dst,(3,3),0)

    _,thresh = cv2.threshold(blur,230,255,cv2.THRESH_BINARY )
    #thresh = cv2.erode(thresh,(9,9),iterations = 1 )
    thresh = cv2.erode(thresh,(5,5),iterations = 1 )

    #thresh[:,2485:] = 0
    thresh[:,2460:] = 0

    thresh[:,0:100] = 0

    return thresh

def createNewFile(csvName):
    with open(csvName,"w+") as f:
        #print("3")
        writer = csv.writer(f)
        f.writelines("test")
        f.close()
        #print("---------------------------------------")
        #time.sleep(5)
        #print("4")
            
def writeCSV(valCsv):
                global ui
                csvName = "thu_28_may.csv"
                try:
                    #print("1")
                    df = pd.read_csv(csvName,header = [0])
                    #print(df)
                except FileNotFoundError:
                    #print("2")
                    createNewFile(csvName)
                    #print("5")
                    df = pd.read_csv(csvName,header=[0],low_memory=False)
                #print("6")
                valCsv[0] = time.strftime("%Y-%m-%d(%H_%M_%S)")
                df1 = pd.DataFrame(valCsv)
                df = pd.concat([df ,df1], ignore_index=True, axis=1)
                df.to_csv(csvName,index=False)
                df1 = None
                df = None
                print("ask to clear graph---------------------")
                ui.clearGraph()



def slabDim(dimCameraLiink):

    global lookForOcr
    #print(lookForOcr)
    global timeOver
    global valCsv
    global ui
    global Ui_MainWindow
    
    lookForOcr = False
    timeOver = 0
    #alarRef = [410,1165,2165,1165]
    alarRef = [365,1165,2200,1165] # 1_june_2020(10.51_am)
    #alar = [alarRef[0],alarRef[1]-40,alarRef[2],alarRef[3]-40]
    alar = [alarRef[0],alarRef[1]-80,alarRef[2],alarRef[3]-80]
    
    #alarRef = [330,820,1585,820] # trail_ggjh
    #alar = [alarRef[0],alarRef[1]-40,alarRef[2],alarRef[3]-40] #  # trail_ggjh

    dB = dist.euclidean(alarRef[:2],alarRef[-2:])
    #dB = alar[2]-alar[0]
    #print(str(dB)+'====sub')
    pixelsPerMetric = dB / 1690 #(pixel/mm)
    #print(str(pixelsPerMetric)+'***pixelPerMetric')
    print(str(1/pixelsPerMetric)+'--------mmPerPixel')

    valCsv = []
    dim = cv2.VideoCapture(dimCameraLiink)
    #dim = cv2.VideoCapture("C:/Users/CCTV/Desktop/Videos/1280.mov")
    dim.set(cv2.CAP_PROP_BUFFERSIZE, 10)
    #dim = cv2.VideoCapture('rtsp://admin:TATA_tsk123@10.152.235.180')
    #dim.set(cv2.CAP_PROP_POS_FRAMES,1)
    
    while(True):
        #st1 = time.time()

        ret, img_dim = dim.read()

        val = Ui_MainWindow.retunValue()
        alarRef = [365,1165+val,2200,1165+val] # 05_june_2020(12.20_pm)
        alar = [alarRef[0],alarRef[1]-80,alarRef[2],alarRef[3]-80]

        #img_dim[:,2485:] = 0
        #img_dim[:,0:100] = 0
        if ret == False:
            #print((time.time()-st1)*1000,"ms---")
            print("Slab camera Video could not be fetched --- Kindly re-run the application")
            print(time.strftime("%Y-%m-%d(%H_%M_%S)"))
            dim = cv2.VideoCapture(dimCameraLiink)
            #dim.set(cv2.CAP_PROP_BUFFERSIZE, 10)
            continue
        #img_dim = imutils.rotate_bound(img_dim,.5)
        thresh = threshOP(img_dim) #----------------------------------------------------------------------
        #if np.count_nonzero(thresh[alar[1],]) >= 550: # traila_igjyg        
        if np.count_nonzero(thresh[alar[1],]) >= 950 and lookForOcr == False:            
            mmPPix,img_dim = mmPerPix(img_dim,thresh,alar,pixelsPerMetric) #-----------------------------------------------------
            valCsv.append(mmPPix)
            '''if len(valCsv) > 5 and len(valCsv) < 150:
                t3 = threading.Thread(target=ui.updateGraph,args=(valCsv,))
                t3.start()
                t3.join()
                if (len(valCsv) % 6) == 0:
                    ui.updateGraph(valCsv)
                    dim = cv2.VideoCapture(dimCameraLiink)'''
            
            # considaring we have a longest slab
            if len(valCsv)> 150:
                if lookForOcr == False:
                    lookForOcr = True
                    timeOver = time.time()
                    print('----------------------------------------------max value reached >150')
                    print(lookForOcr,"<-look for ocr")
                    ui.updateGraph(valCsv)
                    #dim = cv2.VideoCapture(dimCameraLiink)
                    #dim.set(cv2.CAP_PROP_BUFFERSIZE, 10)
        # slab passes the ROI
        #if np.count_nonzero(thresh[alar[1]+30,]) <= 550 and np.count_nonzero(thresh[alar[1]-30,]) <= 550: # trail_jkghjvjhg
        if np.count_nonzero(thresh[alar[1]+30,]) <= 950 and np.count_nonzero(thresh[alar[1]-30,]) <= 950:
            if len(valCsv)> 60:
                if lookForOcr == False:
                    lookForOcr = True
                    timeOver = time.time()
                    #writeCSV(valCsv)
                    #valCsv = []
                    print("+++++++++++++++++++++++++ slab over")
                    print(lookForOcr,"<-look for ocr")
                    ui.updateGraph(valCsv)
                    #dim = cv2.VideoCapture(dimCameraLiink)
                    #dim.set(cv2.CAP_PROP_BUFFERSIZE, 10)

            else:
                None
                #valCsv = []
        # Ref line
        #img_dim  = cv2.line(img_dim ,(alar[0] ,alar[1]+60),(alar[2] ,alar[3]+60),(250,0,0),4) # ref line
        img_dim  = cv2.line(img_dim ,(alarRef[0] ,alarRef[1]),(alarRef[2] ,alarRef[3]),(255,0,0),4) # ref line
        img_dim = cv2.resize(img_dim ,(int(1200),int(700)))

        img_dim = cv2.cvtColor(img_dim, cv2.COLOR_BGR2RGB)
        img_dim = np.require(img_dim,np.uint8,'C')
        ui.showImg1(img_dim)
        
        #dim = cv2.VideoCapture(dimCameraLiink)
        #dim.set(cv2.CAP_PROP_BUFFERSIZE, 10)
        #continue
        # Resize
        #img_dim1 = cv2.resize(img_dim ,(int(1200),int(700)))
        #thresh1 = cv2.resize(thresh,(int(1920/2),int(1080/2)))
        
        #cv2.imshow('img_dim ',img_dim1)
        #cv2.imshow(' thrsh ',thresh1)
        #key = cv2.waitKey(1) & 0xFF
        
        #print((time.time()-st1)*1000,"ms---")
        '''if key == ord('q'):
            break


    dim.release()
    cv2.destroyAllWindows()'''


'''def startThread():
    dimCameraLiink = "rtsp://admin:TATA_tsk123@10.152.235.180"
    ocrCameraLink = "rtsp://service:TATA_tsk123@10.152.235.184/ch1-s1?tcp" #10.152.235.184
    
    t1 = threading.Thread(target=slabDim,args=(dimCameraLiink,))
    t2 = threading.Thread(target=ocr,args=(ocrCameraLink,))
    
    t1.start()
    t2.start()

    
    t1.join()
    t2.join()'''
    


def main():
    global ui
    global MainWindow
    global Ui_MainWindow

    pid = win32api.GetCurrentProcessId()
    print(pid,"<----Process ID")
    handle = win32api.OpenProcess(win32con.PROCESS_ALL_ACCESS, True, pid)
    win32process.SetPriorityClass(handle, win32process.HIGH_PRIORITY_CLASS)

    
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()

      
    dimCameraLiink = "rtsp://admin:TATA_tsk123@10.152.235.180"
    ocrCameraLink = "rtsp://service:TATA_tsk123@10.152.235.184/ch1-s1?tcp" #10.152.235.184
    
    t1 = threading.Thread(target=slabDim,args=(dimCameraLiink,))
    t2 = threading.Thread(target=ocr,args=(ocrCameraLink,))
    #t3 = threading.Thread(target = uiocrView,args=(ocrCameraLink,))
    
    t1.start()
    t2.start()
    #t3.start()
    sys.exit(app.exec_())

    
    t1.join()
    t2.join()
    #t3.join()
    print("Done with both the Threads................")

if __name__ == '__main__':
    main()
